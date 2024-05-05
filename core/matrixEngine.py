from dotenv import load_dotenv
from langchain_openai import AzureChatOpenAI
from langchain_core.pydantic_v1 import BaseModel, Field
from langchain.prompts import PromptTemplate
from langchain.output_parsers.openai_tools import JsonOutputToolsParser

from core.config import AZURE_OPENAI_ENDPOINT,AZURE_OPENAI_DEPLOYMENT,AZURE_OPENAI_MODEL
import openpyxl
import sqlite3
import json


class CallClassification(BaseModel):
    summary: str = Field(description="Summary of the call transcript.")
    sentiment: str = Field(description="Sentiment of the call. It should be one of the following: positive, negative, neutral")
    emotion: str = Field(description="Emotion of the call transcript.")
    category: str = Field(description="Category of the call transcript. It should be one of the following:  ['New Customer Inquiry', 'Billing Inquiry', 'Technical Support', 'Service Upgrade or Add-Ons', 'Account Cancellation']")

class MatrixGPT:
    def add_to_DB(cursor,filepath):
        #CREATING PROMPT TEMPLATE
        template = """
        You are a professional customer call classifier. You are given a call transcript between a customer and an agent representing MatrixTel, A Telecom Company. Your task is to classify the call transcript into one of the following categories: ['New Customer Inquiry', 'Billing Inquiry', 'Technical Support', 'Service Upgrade or Add-Ons', 'Account Cancellation'], and also provide the sentiment and emotion of the call transcript.

        call transcript:
        {call_transcript}
        """
        llm = AzureChatOpenAI(deployment_name=AZURE_OPENAI_DEPLOYMENT, model_name=AZURE_OPENAI_MODEL, temperature=0)
        prompt = PromptTemplate.from_template(template)


        #CREATING LLM CHAIN
        chain = prompt | llm.bind_tools([CallClassification]) | JsonOutputToolsParser()


        wrkbk = openpyxl.load_workbook(filepath) 
        sh = wrkbk.active 
        # iterate through excel and display data 
        for i in range(2, sh.max_row+1): 
            print("\n") 
            print("Row ", i, " data :") 
            cell_obj = sh.cell(row=i, column=2) 
            call_transcript=cell_obj.value
             
            res = chain.invoke({"call_transcript": call_transcript, "question": "What is the summary, sentiment, emotion and category of the call transcript?"})
            data = res[0]['args']
            summary=data['summary']
            sentiment=data['sentiment']
            emotion=data['emotion']
            category=data['category']
            cell_obj = sh.cell(row=i, column=3)
            call_date = cell_obj.value
            cell_obj = sh.cell(row=i, column=4)
            call_duration = cell_obj.value
            print(summary +' '+category)
            
            transcript=(i,call_transcript,call_date,call_duration,summary,sentiment,emotion,category)
            sql= '''INSERT INTO transcript(id,CALL_DETAILS,DATE_OF_CALL,DURATION,summary,sentiment,emotion,category)values(?,?,?,?,?,?,?,?)'''
            cursor.execute(sql, transcript)
            
    def generateInsights(call_transcript):
        template = """
        You are a professional customer call classifier. You are given a call transcript between a customer and an agent representing MatrixTel, A Telecom Company. Your task is to classify the call transcript into one of the following categories: ['New Customer Inquiry', 'Billing Inquiry', 'Technical Support', 'Service Upgrade or Add-Ons', 'Account Cancellation'], and also provide the sentiment and emotion of the call transcript.

        call transcript:
        {call_transcript}
        """
        llm = AzureChatOpenAI(deployment_name=AZURE_OPENAI_DEPLOYMENT, model_name=AZURE_OPENAI_MODEL, temperature=0)
        prompt = PromptTemplate.from_template(template)


        #CREATING LLM CHAIN
        chain = prompt | llm.bind_tools([CallClassification]) | JsonOutputToolsParser()
        res = chain.invoke({"call_transcript": call_transcript, "question": "What is the summary, sentiment, emotion and category of the call transcript?"})
        data = res[0]['args']
        summary=data['summary']
        sentiment=data['sentiment']
        emotion=data['emotion']
        category=data['category']
        
        return data